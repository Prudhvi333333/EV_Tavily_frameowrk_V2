from __future__ import annotations

import html
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.utils.config_loader import resolve_path


METRIC_COLUMN_MAP = {
    "faithfulness": "Faithfulness",
    "answer_relevancy": "Answer_Relevancy",
    "context_precision": "Context_Precision",
    "context_recall": "Context_Recall",
    "answer_correctness": "Answer_Correctness",
    "source_attribution": "Source_Attribution",
    "web_grounding": "Web_Grounding",
}


def build_report(
    results: list[dict[str, Any]],
    model_name: str,
    pipeline_mode: str,
    config: dict[str, Any],
    train_results: list[dict[str, Any]] | None = None,
) -> str:
    reports_dir = resolve_path(config, config["paths"]["reports_dir"])
    reports_dir.mkdir(parents=True, exist_ok=True)
    report_path = Path(reports_dir) / f"{model_name}_{pipeline_mode}_report.xlsx"

    rows = [_row_to_result_record(r, pipeline_mode) for r in results]
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame([{"Note": "No results available."}])

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        _build_summary_sheet(df).to_excel(writer, sheet_name="Summary", index=False)
        _build_train_test_sheet(df, train_results).to_excel(writer, sheet_name="Train_vs_Test", index=False)
        _build_validation_audit_sheet(results).to_excel(writer, sheet_name="Validation_Audit", index=False)
        if pipeline_mode == "rag_pretrained_web":
            _build_web_validation_sheet(results).to_excel(writer, sheet_name="Web_Validation", index=False)

    _apply_conditional_colors(report_path, sheet_name="Results")
    return str(report_path)


def build_comparison_report(all_results: dict[tuple[str, str], list[dict[str, Any]]], config: dict[str, Any]) -> str:
    reports_dir = resolve_path(config, config["paths"]["reports_dir"])
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = Path(reports_dir) / "FINAL_COMPARISON.xlsx"

    agg_rows: list[dict[str, Any]] = []
    for (model, pipeline), rows in all_results.items():
        metric_matrix = pd.DataFrame([r.get("metric_scores", {}) for r in rows])
        means = metric_matrix.mean(numeric_only=True).to_dict() if not metric_matrix.empty else {}
        agg_rows.append(
            {
                "Model": model,
                "Pipeline": pipeline,
                "Final_Score_Mean": float(np.mean([r.get("final_score", 0.0) for r in rows])) if rows else 0.0,
                **{METRIC_COLUMN_MAP.get(k, k): float(v) for k, v in means.items()},
            }
        )
    agg_df = pd.DataFrame(agg_rows).sort_values(["Model", "Pipeline"])

    rag_vs_no = _build_rag_vs_norag(agg_df)
    rankings = agg_df.sort_values("Final_Score_Mean", ascending=False).reset_index(drop=True)
    pretrained_delta = _build_delta_sheet(agg_df, base_pipeline="rag", compare_pipeline="rag_pretrained", label="Pretrained")
    web_delta = _build_delta_sheet(agg_df, base_pipeline="rag_pretrained", compare_pipeline="rag_pretrained_web", label="Web")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        agg_df.to_excel(writer, sheet_name="All_12_Pipelines", index=False)
        rag_vs_no.to_excel(writer, sheet_name="RAG_vs_NoRAG", index=False)
        rankings.to_excel(writer, sheet_name="Rankings", index=False)
        pretrained_delta.to_excel(writer, sheet_name="Pretrained_Contribution", index=False)
        web_delta.to_excel(writer, sheet_name="Web_Contribution", index=False)

    return str(out_path)


def build_reviewer_dashboard(
    all_results: dict[tuple[str, str], list[dict[str, Any]]],
    config: dict[str, Any],
    progress_runs: list[dict[str, Any]] | None = None,
    comparison_report_path: str | None = None,
    proof_since_utc: datetime | None = None,
) -> str:
    reports_dir = resolve_path(config, config["paths"]["reports_dir"])
    reports_dir.mkdir(parents=True, exist_ok=True)
    dashboard_path = Path(reports_dir) / "REVIEWER_DASHBOARD.html"

    question_rows: list[dict[str, Any]] = []
    web_rows: list[dict[str, Any]] = []
    for (model, pipeline), rows in all_results.items():
        for row in rows:
            final_score = row.get("final_score", np.nan)
            question_rows.append(
                {
                    "q_id": row.get("q_id", ""),
                    "model": model,
                    "pipeline": pipeline,
                    "question": row.get("question", ""),
                    "web_status": row.get("web_status", "NOT_USED"),
                    "final_score": final_score,
                    "validation_flags": row.get("validation_flags", ""),
                    "validation_reason": row.get("validation_reason", ""),
                }
            )
            for rec in row.get("web_validation_records", []):
                signals = rec.get("signals", {})
                web_rows.append(
                    {
                        "q_id": row.get("q_id", ""),
                        "model": model,
                        "pipeline": pipeline,
                        "url": rec.get("url", ""),
                        "source_domain": rec.get("source_domain", ""),
                        "decision": rec.get("decision", ""),
                        "accepted": bool(rec.get("accepted", False)),
                        "low_confidence": bool(rec.get("low_confidence", False)),
                        "final_score": rec.get("final_score", np.nan),
                        "s1": signals.get("s1_keyword", np.nan),
                        "s2": signals.get("s2_semantic", np.nan),
                        "s3": signals.get("s3_llm", np.nan),
                        "reason": rec.get("s3_reason", ""),
                        "preview": rec.get("text_preview", ""),
                    }
                )

    mean_final = float(np.nanmean([_safe_float(r.get("final_score", np.nan)) for r in question_rows])) if question_rows else np.nan
    accepted_count = sum(1 for r in web_rows if r["accepted"] and not r["low_confidence"])
    low_confidence_count = sum(1 for r in web_rows if r["accepted"] and r["low_confidence"])
    rejected_count = sum(1 for r in web_rows if not r["accepted"])

    all_proof_rows = _load_recent_proof_rows(config, limit=300)
    if proof_since_utc is not None:
        filtered_rows: list[dict[str, Any]] = []
        for rec in all_proof_rows:
            rec_ts = _parse_iso_utc(rec.get("timestamp", ""))
            if rec_ts and rec_ts >= proof_since_utc:
                filtered_rows.append(rec)
        all_proof_rows = filtered_rows
    active_proof_keys = {
        (str(row.get("question", "")).strip(), f"{row.get('model', '')}_{row.get('pipeline', '')}")
        for row in question_rows
    }
    proof_rows = [
        rec
        for rec in all_proof_rows
        if (str(rec.get("question", "")).strip(), str(rec.get("pipeline", "")).strip()) in active_proof_keys
    ]
    if not proof_rows:
        proof_rows = all_proof_rows
    html_report = _render_dashboard_html(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        comparison_report_path=comparison_report_path,
        progress_runs=progress_runs or [],
        question_rows=question_rows,
        web_rows=web_rows,
        proof_rows=proof_rows,
        mean_final=mean_final,
        accepted_count=accepted_count,
        low_confidence_count=low_confidence_count,
        rejected_count=rejected_count,
    )
    dashboard_path.write_text(html_report, encoding="utf-8")
    return str(dashboard_path)


def _row_to_result_record(row: dict[str, Any], pipeline_mode: str) -> dict[str, Any]:
    out = {
        "Q_ID": row.get("q_id", ""),
        "Category": row.get("category", ""),
        "Question": row.get("question", ""),
        "Golden": row.get("golden", ""),
        "Answer": row.get("answer", ""),
    }
    if pipeline_mode != "no_rag":
        out["Context"] = row.get("kb_context", "")
    if pipeline_mode == "rag_pretrained_web":
        out["Web_Context"] = row.get("web_context", "")
        out["Web_Status"] = row.get("web_status", "")
        out["Web_Search_Query"] = row.get("web_search_query", "")
        out["Web_Timed_Out"] = bool(row.get("web_timed_out", False))
        out["Web_Accepted_Docs"] = row.get("web_accepted_count", 0)
        out["Web_Low_Confidence_Docs"] = row.get("web_low_confidence_count", 0)
        out["Web_Rejected_Docs"] = row.get("web_rejected_count", 0)

    metric_scores = row.get("metric_scores", {})
    for metric, value in metric_scores.items():
        out[METRIC_COLUMN_MAP.get(metric, metric)] = value

    if pipeline_mode in {"rag_pretrained", "rag_pretrained_web"}:
        answer_text = str(row.get("answer", ""))
        pretrained_tags = answer_text.count("[PRETRAINED]")
        kb_tags = max(answer_text.count("[KB]"), 1)
        out["Pretrained_Leak_Rate"] = round(pretrained_tags / kb_tags, 4)

    out["Final_Score"] = row.get("final_score", np.nan)
    out["Validation_Flags"] = row.get("validation_flags", "")
    out["Validation_Reason"] = row.get("validation_reason", "")
    return out


def _build_web_validation_sheet(results: list[dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for result in results:
        q_id = result.get("q_id", "")
        question = result.get("question", "")
        model_key = result.get("model_key", "")
        pipeline_mode = result.get("pipeline_mode", "")
        for record in result.get("web_validation_records", []):
            signals = record.get("signals", {})
            rows.append(
                {
                    "Q_ID": q_id,
                    "Question": question,
                    "Model": model_key,
                    "Pipeline": pipeline_mode,
                    "URL": record.get("url", ""),
                    "Source_Domain": record.get("source_domain", ""),
                    "Decision": record.get("decision", ""),
                    "Accepted": bool(record.get("accepted", False)),
                    "Low_Confidence": bool(record.get("low_confidence", False)),
                    "Final_Score": record.get("final_score", np.nan),
                    "S1_Keyword": signals.get("s1_keyword", np.nan),
                    "S2_Semantic": signals.get("s2_semantic", np.nan),
                    "S3_LLM": signals.get("s3_llm", np.nan),
                    "S3_Reason": record.get("s3_reason", ""),
                    "Text_Preview": record.get("text_preview", ""),
                }
            )

    if not rows:
        rows = [
            {
                "Q_ID": "",
                "Question": "No web validation records available for this run.",
                "Model": "",
                "Pipeline": "",
                "URL": "",
                "Source_Domain": "",
                "Decision": "",
                "Accepted": "",
                "Low_Confidence": "",
                "Final_Score": np.nan,
                "S1_Keyword": np.nan,
                "S2_Semantic": np.nan,
                "S3_LLM": np.nan,
                "S3_Reason": "",
                "Text_Preview": "",
            }
        ]
    return pd.DataFrame(rows)


def _build_summary_sheet(df: pd.DataFrame) -> pd.DataFrame:
    metric_cols = [c for c in df.columns if c in set(METRIC_COLUMN_MAP.values()) | {"Final_Score"}]
    if not metric_cols:
        return pd.DataFrame([{"Metric": "N/A", "Mean": np.nan, "Std": np.nan}])
    rows = []
    for col in metric_cols:
        rows.append({"Metric": col, "Mean": float(df[col].mean()), "Std": float(df[col].std(ddof=0))})
    return pd.DataFrame(rows)


def _build_train_test_sheet(test_df: pd.DataFrame, train_results: list[dict[str, Any]] | None) -> pd.DataFrame:
    if not train_results:
        return pd.DataFrame(
            [{"Split": "test", "Final_Score_Mean": float(test_df["Final_Score"].mean()) if "Final_Score" in test_df else np.nan}]
        )
    train_scores = [float(r.get("final_score", 0.0)) for r in train_results]
    test_scores = test_df["Final_Score"].dropna().tolist() if "Final_Score" in test_df else []
    return pd.DataFrame(
        [
            {"Split": "train", "Final_Score_Mean": float(np.mean(train_scores)) if train_scores else np.nan},
            {"Split": "test", "Final_Score_Mean": float(np.mean(test_scores)) if test_scores else np.nan},
        ]
    )


def _build_validation_audit_sheet(results: list[dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for r in results:
        if not r.get("validation_flags"):
            continue
        rows.append(
            {
                "Q_ID": r.get("q_id", ""),
                "Question": r.get("question", ""),
                "Flagged_Metrics": r.get("flagged_metrics", ""),
                "Reason": r.get("validation_reason", ""),
                "Adjustments": str(r.get("metric_adjustments", {})),
            }
        )
    if not rows:
        rows = [{"Q_ID": "", "Question": "No flagged rows", "Flagged_Metrics": "", "Reason": "", "Adjustments": ""}]
    return pd.DataFrame(rows)


def _apply_conditional_colors(report_path: str | Path, sheet_name: str) -> None:
    wb = load_workbook(report_path)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    metric_columns = {"Faithfulness", "Answer_Relevancy", "Context_Precision", "Context_Recall", "Answer_Correctness", "Source_Attribution", "Web_Grounding", "Final_Score"}
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        if header not in metric_columns:
            continue
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, (int, float)):
                if val >= 0.7:
                    ws.cell(row=row_idx, column=col_idx).fill = green
                elif val >= 0.5:
                    ws.cell(row=row_idx, column=col_idx).fill = yellow
                else:
                    ws.cell(row=row_idx, column=col_idx).fill = red

    wb.save(report_path)


def _build_rag_vs_norag(agg_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for model in sorted(agg_df["Model"].unique()):
        mdf = agg_df[agg_df["Model"] == model]
        rag = mdf[mdf["Pipeline"] == "rag"]
        no_rag = mdf[mdf["Pipeline"] == "no_rag"]
        rows.append(
            {
                "Model": model,
                "RAG_Answer_Correctness": float(rag["Answer_Correctness"].iloc[0]) if not rag.empty and "Answer_Correctness" in rag else np.nan,
                "NoRAG_Answer_Correctness": float(no_rag["Answer_Correctness"].iloc[0]) if not no_rag.empty and "Answer_Correctness" in no_rag else np.nan,
            }
        )
    return pd.DataFrame(rows)


def _build_delta_sheet(agg_df: pd.DataFrame, base_pipeline: str, compare_pipeline: str, label: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for model in sorted(agg_df["Model"].unique()):
        mdf = agg_df[agg_df["Model"] == model]
        base = mdf[mdf["Pipeline"] == base_pipeline]
        comp = mdf[mdf["Pipeline"] == compare_pipeline]
        if base.empty or comp.empty:
            delta = np.nan
        else:
            delta = float(comp["Final_Score_Mean"].iloc[0] - base["Final_Score_Mean"].iloc[0])
        rows.append({"Model": model, f"{label}_Delta": delta})
    return pd.DataFrame(rows)


def _safe_float(value: Any) -> float:
    try:
        val = float(value)
        if np.isnan(val):
            return np.nan
        return val
    except Exception:
        return np.nan


def _fmt_score(value: Any) -> str:
    val = _safe_float(value)
    if np.isnan(val):
        return "N/A"
    return f"{val:.3f}"


def _score_class(value: Any) -> str:
    val = _safe_float(value)
    if np.isnan(val):
        return "score-na"
    if val >= 0.7:
        return "score-high"
    if val >= 0.5:
        return "score-mid"
    return "score-low"


def _clip_text(text: Any, limit: int) -> str:
    raw = str(text or "").strip()
    if len(raw) <= limit:
        return raw
    return raw[: limit - 3] + "..."


def _parse_iso_utc(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone(timezone.utc)
    except Exception:
        return None


def _load_recent_proof_rows(config: dict[str, Any], limit: int = 300) -> list[dict[str, Any]]:
    path = resolve_path(config, config["paths"]["logs_dir"]) / "web_validation_proof.jsonl"
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8").splitlines()
    selected = lines[-limit:]
    rows: list[dict[str, Any]] = []
    for line in reversed(selected):
        try:
            rows.append(json.loads(line))
        except Exception:
            continue
    return rows


def _render_dashboard_html(
    generated_at: str,
    comparison_report_path: str | None,
    progress_runs: list[dict[str, Any]],
    question_rows: list[dict[str, Any]],
    web_rows: list[dict[str, Any]],
    proof_rows: list[dict[str, Any]],
    mean_final: float,
    accepted_count: int,
    low_confidence_count: int,
    rejected_count: int,
) -> str:
    comparison_label = html.escape(str(comparison_report_path)) if comparison_report_path else "N/A"
    run_rows_html = "".join(
        f"""
        <tr>
          <td>{html.escape(str(run.get("model", "")))}</td>
          <td>{html.escape(str(run.get("pipeline", "")))}</td>
          <td>{html.escape(str(run.get("questions", 0)))}</td>
          <td><span class="score-chip {_score_class(run.get("mean_final_score", np.nan))}">{_fmt_score(run.get("mean_final_score", np.nan))}</span></td>
          <td>{html.escape(str(run.get("report", "")))}</td>
        </tr>
        """
        for run in progress_runs
    )
    if not run_rows_html:
        run_rows_html = """
        <tr><td colspan="5" class="muted">No run summary rows available.</td></tr>
        """

    question_rows_html = "".join(
        f"""
        <tr>
          <td>{html.escape(str(row.get("q_id", "")))}</td>
          <td>{html.escape(str(row.get("model", "")))}</td>
          <td>{html.escape(str(row.get("pipeline", "")))}</td>
          <td><span class="score-chip {_score_class(row.get("final_score", np.nan))}">{_fmt_score(row.get("final_score", np.nan))}</span></td>
          <td>{html.escape(str(row.get("web_status", "")))}</td>
          <td title="{html.escape(str(row.get("validation_reason", "")))}">{html.escape(str(row.get("validation_flags", "")))}</td>
          <td title="{html.escape(str(row.get("question", "")))}">{html.escape(_clip_text(row.get("question", ""), 120))}</td>
        </tr>
        """
        for row in question_rows
    )
    if not question_rows_html:
        question_rows_html = """
        <tr><td colspan="7" class="muted">No question-level rows available.</td></tr>
        """

    web_rows_html = "".join(
        f"""
        <tr>
          <td>{html.escape(str(row.get("q_id", "")))}</td>
          <td>{html.escape(str(row.get("model", "")))}</td>
          <td>{html.escape(str(row.get("pipeline", "")))}</td>
          <td>{html.escape(str(row.get("source_domain", "")))}</td>
          <td><a href="{html.escape(str(row.get("url", "")))}" target="_blank" rel="noopener noreferrer">open</a></td>
          <td>{html.escape(str(row.get("decision", "")))}</td>
          <td><span class="score-chip {_score_class(row.get("final_score", np.nan))}">{_fmt_score(row.get("final_score", np.nan))}</span></td>
          <td>{_fmt_score(row.get("s1", np.nan))}</td>
          <td>{_fmt_score(row.get("s2", np.nan))}</td>
          <td>{_fmt_score(row.get("s3", np.nan))}</td>
          <td title="{html.escape(str(row.get("reason", "")))}">{html.escape(_clip_text(row.get("reason", ""), 80))}</td>
          <td title="{html.escape(str(row.get("preview", "")))}">{html.escape(_clip_text(row.get("preview", ""), 120))}</td>
        </tr>
        """
        for row in web_rows
    )
    if not web_rows_html:
        web_rows_html = """
        <tr><td colspan="12" class="muted">No web validation records in this run.</td></tr>
        """

    proof_rows_html = "".join(
        f"""
        <tr>
          <td>{html.escape(str(row.get("timestamp", "")))}</td>
          <td>{html.escape(str(row.get("question_id", "")))}</td>
          <td>{html.escape(str(row.get("pipeline", "")))}</td>
          <td>{html.escape(str(row.get("source_domain", "")))}</td>
          <td>{html.escape(str(row.get("decision", "")))}</td>
          <td><span class="score-chip {_score_class(row.get("final_score", np.nan))}">{_fmt_score(row.get("final_score", np.nan))}</span></td>
          <td title="{html.escape(str(row.get("s3_reason", "")))}">{html.escape(_clip_text(row.get("s3_reason", ""), 80))}</td>
          <td title="{html.escape(str(row.get("text_preview", "")))}">{html.escape(_clip_text(row.get("text_preview", ""), 120))}</td>
        </tr>
        """
        for row in proof_rows
    )
    if not proof_rows_html:
        proof_rows_html = """
        <tr><td colspan="8" class="muted">No proof log entries found at outputs/logs/web_validation_proof.jsonl.</td></tr>
        """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>RAG Reviewer Dashboard</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&family=Source+Serif+4:wght@400;600&display=swap');
    :root {{
      --bg: #f3f6ef;
      --panel: #ffffff;
      --ink: #122117;
      --muted: #5d6b60;
      --line: #d7e1d2;
      --high: #266b36;
      --mid: #8a6d16;
      --low: #8f2d20;
      --accent: #0f766e;
      --accent-soft: #d9f3f0;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: 'Space Grotesk', sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at 15% -10%, #d7efe6 0, transparent 40%),
        radial-gradient(circle at 90% 0, #f8e5c9 0, transparent 32%),
        var(--bg);
      line-height: 1.4;
    }}
    .wrap {{
      max-width: 1320px;
      margin: 0 auto;
      padding: 24px 16px 40px;
    }}
    .hero {{
      background: linear-gradient(120deg, #08352f, #185f54);
      color: #f3fff9;
      border-radius: 16px;
      padding: 18px 20px;
      box-shadow: 0 10px 28px rgba(6, 40, 34, 0.2);
    }}
    .hero h1 {{
      margin: 0 0 8px;
      font-size: 28px;
      letter-spacing: 0.3px;
    }}
    .hero p {{
      margin: 2px 0;
      color: #d4efe8;
      font-family: 'Source Serif 4', serif;
    }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
      gap: 12px;
      margin: 16px 0 20px;
    }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 12px 14px;
      box-shadow: 0 4px 14px rgba(28, 56, 41, 0.06);
    }}
    .card .label {{
      font-size: 12px;
      letter-spacing: 0.5px;
      text-transform: uppercase;
      color: var(--muted);
    }}
    .card .value {{
      margin-top: 4px;
      font-size: 24px;
      font-weight: 700;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 12px;
      margin-top: 12px;
      box-shadow: 0 4px 14px rgba(28, 56, 41, 0.06);
    }}
    .panel h2 {{
      margin: 0 0 8px;
      font-size: 18px;
    }}
    .panel .sub {{
      margin: 0 0 10px;
      color: var(--muted);
      font-size: 13px;
    }}
    .toolbar {{
      margin: 8px 0 10px;
    }}
    .toolbar input {{
      width: 100%;
      max-width: 460px;
      border: 1px solid var(--line);
      background: #f9fcf7;
      border-radius: 10px;
      padding: 8px 10px;
      font: inherit;
      color: var(--ink);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    th, td {{
      border-bottom: 1px solid #edf2eb;
      text-align: left;
      vertical-align: top;
      padding: 8px 8px;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #f7fbf5;
      color: #304438;
      z-index: 1;
    }}
    .table-wrap {{
      max-height: 360px;
      overflow: auto;
      border: 1px solid #eef3ec;
      border-radius: 8px;
    }}
    .score-chip {{
      display: inline-block;
      min-width: 56px;
      text-align: center;
      border-radius: 999px;
      padding: 2px 8px;
      font-weight: 700;
    }}
    .score-high {{ background: #d8f2df; color: var(--high); }}
    .score-mid {{ background: #fff2ca; color: var(--mid); }}
    .score-low {{ background: #ffdcd5; color: var(--low); }}
    .score-na {{ background: #eceff1; color: #54606b; }}
    .muted {{ color: var(--muted); }}
    a {{ color: var(--accent); text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>RAG Evaluation Reviewer Dashboard</h1>
      <p>Generated: {html.escape(generated_at)}</p>
      <p>Comparison Workbook: {comparison_label}</p>
    </section>

    <section class="cards">
      <article class="card"><div class="label">Run Rows</div><div class="value">{len(progress_runs)}</div></article>
      <article class="card"><div class="label">Questions</div><div class="value">{len(question_rows)}</div></article>
      <article class="card"><div class="label">Mean Final Score</div><div class="value">{_fmt_score(mean_final)}</div></article>
      <article class="card"><div class="label">Web Accepted</div><div class="value">{accepted_count}</div></article>
      <article class="card"><div class="label">Low Confidence</div><div class="value">{low_confidence_count}</div></article>
      <article class="card"><div class="label">Web Rejected</div><div class="value">{rejected_count}</div></article>
      <article class="card"><div class="label">Proof Entries</div><div class="value">{len(proof_rows)}</div></article>
    </section>

    <section class="panel">
      <h2>Run Summary</h2>
      <p class="sub">High-level results per model and pipeline run.</p>
      <div class="table-wrap">
        <table id="run-table">
          <thead><tr><th>Model</th><th>Pipeline</th><th>Questions</th><th>Mean Score</th><th>Report Path</th></tr></thead>
          <tbody>{run_rows_html}</tbody>
        </table>
      </div>
    </section>

    <section class="panel">
      <h2>Question Outcomes</h2>
      <p class="sub">Per-question scores, flags, and web status.</p>
      <div class="toolbar"><input id="question-filter" type="text" placeholder="Filter question rows..." /></div>
      <div class="table-wrap">
        <table id="question-table">
          <thead><tr><th>Q_ID</th><th>Model</th><th>Pipeline</th><th>Final</th><th>Web Status</th><th>Flags</th><th>Question</th></tr></thead>
          <tbody>{question_rows_html}</tbody>
        </table>
      </div>
    </section>

    <section class="panel">
      <h2>Web Validation</h2>
      <p class="sub">Signal-level validation for each crawled document.</p>
      <div class="toolbar"><input id="web-filter" type="text" placeholder="Filter web validation rows..." /></div>
      <div class="table-wrap">
        <table id="web-table">
          <thead><tr><th>Q_ID</th><th>Model</th><th>Pipeline</th><th>Domain</th><th>URL</th><th>Decision</th><th>Final</th><th>S1</th><th>S2</th><th>S3</th><th>Judge Reason</th><th>Preview</th></tr></thead>
          <tbody>{web_rows_html}</tbody>
        </table>
      </div>
    </section>

    <section class="panel">
      <h2>Proof Log (Latest 300)</h2>
      <p class="sub">Direct JSONL evidence written during validation.</p>
      <div class="toolbar"><input id="proof-filter" type="text" placeholder="Filter proof rows..." /></div>
      <div class="table-wrap">
        <table id="proof-table">
          <thead><tr><th>Timestamp</th><th>Question ID</th><th>Pipeline</th><th>Domain</th><th>Decision</th><th>Final</th><th>Judge Reason</th><th>Preview</th></tr></thead>
          <tbody>{proof_rows_html}</tbody>
        </table>
      </div>
    </section>
  </div>
  <script>
    function bindFilter(inputId, tableId) {{
      const input = document.getElementById(inputId);
      const table = document.getElementById(tableId);
      if (!input || !table) return;
      const rows = Array.from(table.querySelectorAll("tbody tr"));
      input.addEventListener("input", () => {{
        const q = input.value.toLowerCase().trim();
        rows.forEach((row) => {{
          row.style.display = row.innerText.toLowerCase().includes(q) ? "" : "none";
        }});
      }});
    }}
    bindFilter("question-filter", "question-table");
    bindFilter("web-filter", "web-table");
    bindFilter("proof-filter", "proof-table");
  </script>
</body>
</html>
"""
